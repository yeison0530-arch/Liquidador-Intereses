import streamlit as st
import pandas as pd
import requests
from datetime import datetime, timedelta
import io
from fpdf import FPDF
import re
from PIL import Image
from streamlit_drawable_canvas import st_canvas
import numpy as np
import openpyxl
from openpyxl.styles import Font
from openpyxl.drawing.image import Image as OpenpyxlImage
import tempfile
import os

# Configuración de página
st.set_page_config(page_title="Liquidador de Intereses", layout="wide")

# Diseño Oscuro (CSS inyectado para aproximar la imagen DISEÑO.jpg)
st.markdown("""
<style>
    .stApp { background-color: #16171e; color: #ffffff; }
    h1, h2, h3, h4, h5, span, p { color: #ffffff !important; }
    .stDataFrame { border-radius: 10px; overflow: hidden; }
    .metric-card {
        background-color: #21242d;
        border-radius: 12px;
        padding: 20px;
        margin-bottom: 20px;
        border: 1px solid #333642;
        box-shadow: 0px 4px 10px rgba(0,0,0,0.5);
    }
    .metric-value { font-size: 28px; font-weight: bold; color: #ff4b4b; }
    .metric-label { font-size: 14px; color: #8c8f99; text-transform: uppercase; }
    div.stButton > button {
        background: linear-gradient(90deg, #ff4b4b, #ff7b7b);
        color: white; border: none; border-radius: 8px; font-weight: bold; width: 100%;
        transition: all 0.3s;
    }
    div.stButton > button:hover {
        background: linear-gradient(90deg, #ff7b7b, #ff4b4b);
        color: white; border-color: #ff4b4b; text-shadow: 0 0 5px rgba(255,255,255,0.5);
    }
</style>
""", unsafe_allow_html=True)

st.title("Liquidador de Intereses Moratorios Judiciales")
st.markdown("---")

@st.cache_data(ttl=86400)
def obtener_datos_sfc():
    url = "https://www.datos.gov.co/resource/pare-7x5i.json?$limit=5000"
    response = requests.get(url, headers={'User-Agent': 'Mozilla/5.0'})
    if response.status_code == 200:
        df = pd.DataFrame(response.json())
        df['modalidad'] = df['modalidad'].str.upper()
        df = df[df['modalidad'].str.contains('CONSUMO Y ORDINARIO')]
        df['vigencia_desde'] = pd.to_datetime(df['vigencia_desde'])
        df['vigencia_hasta'] = pd.to_datetime(df['vigencia_hasta'])
        df['interes_bancario_corriente'] = df['interes_bancario_corriente'].str.replace('%','').astype(float)
        return df.sort_values('vigencia_desde').reset_index(drop=True)
    return pd.DataFrame()

df_tasas = obtener_datos_sfc()

def obtener_tasa_vigente(fecha, df_tasas):
    if df_tasas.empty: return 0.0
    fd = pd.Timestamp(fecha)
    mask = (df_tasas['vigencia_desde'] <= fd) & (df_tasas['vigencia_hasta'] >= fd)
    if not mask.any():
        if fd > df_tasas['vigencia_hasta'].max():
            return df_tasas.iloc[-1]['interes_bancario_corriente']
        return 0.0
    return df_tasas.loc[mask, 'interes_bancario_corriente'].values[0]

# --- UI INGRESO DATOS ---
col1, space, col2 = st.columns([1, 0.1, 1])

with col1:
    st.subheader("Cuotas de Capital")
    
    if "cuotas_data" not in st.session_state:
        st.session_state.cuotas_data = pd.DataFrame({
            "Detalle": ["Obligación 1"],
            "Valor Capital": [None],
            "Fecha de Vencimiento": pd.Series([pd.NaT], dtype='datetime64[ns]')
        })

    cuotas_df = st.data_editor(
        st.session_state.cuotas_data,
        num_rows="dynamic",
        key="cuotas_editor",
        hide_index=True,
        use_container_width=True,
        column_config={
            "Detalle": st.column_config.TextColumn("Detalle", help="Ejemplo: 'Pagaré 001' o 'Factura 123'."),
            "Valor Capital": st.column_config.NumberColumn("Valor Capital", format="$ %,.2f", min_value=0.0, help="Escriba el valor numérico sin puntos ni signos adicionales. Ejemplo: Ingrese 1500000.50 y el sistema lo formateará automáticamente como $ 1,500,000.50"),
            "Fecha de Vencimiento": st.column_config.DateColumn("Fecha de Vencimiento", help="Seleccione la fecha exacta en la que debió realizarse el pago.")
        }
    )
    
    if st.button("➕ Añadir Siguiente Cuota Automáticamente"):
        if not cuotas_df.empty:
            ultima_fila = cuotas_df.iloc[-1]
            last_detalle = str(ultima_fila["Detalle"]) if pd.notna(ultima_fila["Detalle"]) and str(ultima_fila["Detalle"]).strip() != "" else ""
            match = re.search(r'(\d+)$', last_detalle)
            if match:
                num = int(match.group(1))
                new_detalle = last_detalle[:match.start()] + str(num + 1)
            else:
                new_detalle = last_detalle + " 2" if last_detalle else "Obligación 2"
                
            new_capital = ultima_fila["Valor Capital"]
            
            new_fecha = pd.NaT
            if pd.notna(ultima_fila["Fecha de Vencimiento"]):
                dt_obj = pd.to_datetime(ultima_fila["Fecha de Vencimiento"])
                new_fecha = dt_obj + pd.DateOffset(months=1)
                
            nueva_fila_df = pd.DataFrame({
                "Detalle": [new_detalle],
                "Valor Capital": [new_capital],
                "Fecha de Vencimiento": pd.Series([new_fecha], dtype='datetime64[ns]')
            })
            
            # Forzar datetime64[ns] sobre cuotas_df antes de concatenar para prevenir el error pyarrow
            cuotas_df["Fecha de Vencimiento"] = pd.to_datetime(cuotas_df["Fecha de Vencimiento"])
            
            st.session_state.cuotas_data = pd.concat([cuotas_df, nueva_fila_df], ignore_index=True)
            
            # Borramos el estado interno del editor para que cargue limpiamente la tabla concatenada
            if "cuotas_editor" in st.session_state:
                del st.session_state["cuotas_editor"]
                
            st.rerun()
    
    st.subheader("Intereses Corrientes Previos")
    st.markdown("<span class='metric-label'>Monto (Valor estático, no anatocismo)</span>", unsafe_allow_html=True)
    int_init = pd.DataFrame(columns=["Detalle", "Monto Interés"])
    int_df = st.data_editor(int_init, num_rows="dynamic", key="intereses", hide_index=True, use_container_width=True,
        column_config={
            "Detalle": st.column_config.TextColumn("Detalle", help="Ejemplo: 'Intereses causados hasta el mes pasado'."),
            "Monto Interés": st.column_config.NumberColumn("Monto Interés", format="$ %,.2f", min_value=0.0, help="Escriba el valor numérico sin puntuación de miles. Ejemplo: Ingrese 150000 y se formateará como $ 150,000.00")
        }
    )
    intereses_previos = float(pd.to_numeric(int_df["Monto Interés"], errors='coerce').sum())

with col2:
    st.subheader("Abonos Realizados")
    abonos_init = pd.DataFrame(columns=["Valor Abono", "Fecha Abono"])
    abonos_df = st.data_editor(abonos_init, num_rows="dynamic", key="abonos", hide_index=True, use_container_width=True,
        column_config={
            "Valor Abono": st.column_config.NumberColumn("Valor Abono", format="$ %,.2f", min_value=0.0, help="Escriba el valor del abono sin puntos de miles. Ejemplo: Ingrese 500000 para referirse a $ 500,000.00"),
            "Fecha Abono": st.column_config.DateColumn("Fecha Abono", help="Seleccione la fecha exacta en la que el comprobante muestra el abono.")
        }
    )
    
    st.subheader("Cortar Liquidación En:")
    fecha_liquidacion = st.date_input("Fecha de Liquidación", value=datetime.today().date(), help="Indica hasta qué fecha se calculará la liquidación de intereses (generalmente la fecha actual).")

# --- OPCIONES DE DOCUMENTO Y FIRMA ---
st.markdown("---")
st.subheader("Opciones del Documento")

texto_inicial = st.text_area("Texto Inicial (Ej. Constancia Secretarial)", help="Este texto se mostrará en la parte superior del documento generado.")

col_f1, col_f2 = st.columns(2)
with col_f1:
    firma_nombre = st.text_input("Nombre completo del firmante", help="Ej. Juan Pérez")
with col_f2:
    firma_cargo = st.text_input("Cargo o Profesión", help="Ej. Secretario")

st.markdown("**Firma Electrónica**")
tipo_firma = st.radio("Método de firma", ["Ninguno (No firma)", "Dibujar con el Mouse", "Subir Imagen"], horizontal=True)

firma_imagen = None
if tipo_firma == "Dibujar con el Mouse":
    st.write("Dibuje su firma en el recuadro blanco:")
    canvas_result = st_canvas(
        fill_color="rgba(255, 165, 0, 0.3)",
        stroke_width=3,
        stroke_color="#000000",
        background_color="#FFFFFF",
        update_streamlit=True,
        height=250,
        width=800,
        drawing_mode="freedraw",
        key="canvas_firma",
    )
    if canvas_result.image_data is not None:
        img_array = canvas_result.image_data
        if np.any(img_array[:, :, 3] > 0): # Check if there is drawn content
            firma_imagen = Image.fromarray(img_array.astype('uint8'), 'RGBA')
elif tipo_firma == "Subir Imagen":
    st.info("💡 **Consejo para pegar:** Selecciona este recuadro y presiona **Ctrl + V** en tu teclado para pegar una firma que hayas copiado de otro lado.")
    archivo_firma = st.file_uploader("Sube o pega una imagen de tu firma", type=["png", "jpg", "jpeg"])
    if archivo_firma is not None:
        firma_imagen = Image.open(archivo_firma)

# --- MATEMÁTICAS Y LÓGICA ---
if st.button("Calcular Liquidación"):
    v_cuotas = cuotas_df.dropna(subset=['Valor Capital', 'Fecha de Vencimiento']).copy()
    v_cuotas = v_cuotas[v_cuotas['Valor Capital'] > 0]
    v_cuotas['Fecha de Vencimiento'] = pd.to_datetime(v_cuotas['Fecha de Vencimiento']).dt.date
    
    v_abonos = abonos_df.dropna(subset=['Valor Abono', 'Fecha Abono']).copy()
    v_abonos = v_abonos[v_abonos['Valor Abono'] > 0]
    v_abonos['Fecha Abono'] = pd.to_datetime(v_abonos['Fecha Abono']).dt.date
    
    if v_cuotas.empty:
        st.error("Ingrese al menos una cuota de capital válida.")
        st.stop()
        
    dates = []
    
    # Cuotas empiezan mora al día siguiente
    for _, r in v_cuotas.iterrows():
        dates.append(r['Fecha de Vencimiento'] + timedelta(days=1))
        
    for _, r in v_abonos.iterrows():
        dates.append(r['Fecha Abono'])
        
    limit_date = fecha_liquidacion + timedelta(days=1)
    dates.append(limit_date)
    
    min_date = min([d for d in dates if d < limit_date]) if len(dates) > 1 else (fecha_liquidacion + timedelta(days=1))  # Fix for edge case
    
    # Generar todos los cortes de mes (día 1)
    cd = (min_date.replace(day=1) + timedelta(days=32)).replace(day=1)
    while cd < limit_date:
        dates.append(cd)
        cd = (cd + timedelta(days=32)).replace(day=1)
        
    # Cambios de tasa
    for td in df_tasas['vigencia_desde'].dt.date:
        if min_date < td < limit_date:
            dates.append(td)
            
    dates = sorted(list(set(dates)))
    dates = [d for d in dates if min_date <= d <= limit_date]
    
    results = []
    capital_base = 0.0
    intereses_acumulados = 0.0
    
    c_list = v_cuotas.to_dict('records')
    a_list = v_abonos.to_dict('records')
    
    for i in range(len(dates)-1):
        start_d = dates[i]
        end_d = dates[i+1]
        
        abono_interes_periodo = 0.0
        abono_capital_periodo = 0.0
        
        # Procesar Capital
        for c in c_list:
            if c['Fecha de Vencimiento'] + timedelta(days=1) == start_d:
                capital_base += c['Valor Capital']
                
        # Procesar Abonos (Art. 1653)
        for a in a_list:
            if a['Fecha Abono'] == start_d:
                m = a['Valor Abono']
                if intereses_acumulados >= m:
                    intereses_acumulados -= m
                    abono_interes_periodo += m
                else:
                    abono_interes_periodo += intereses_acumulados
                    rem = m - intereses_acumulados
                    intereses_acumulados = 0.0
                    capital_base -= rem
                    abono_capital_periodo += rem
                    if capital_base < 0: capital_base = 0.0
                    
        dias = (end_d - start_d).days
        if dias <= 0: continue
        
        ibc = obtener_tasa_vigente(start_d, df_tasas)
        t_mora_anual = (ibc * 1.5) / 100.0
        
        int_gen = capital_base * ((1.0 + t_mora_anual) ** (dias / 365.0) - 1.0)
        intereses_acumulados += int_gen
        
        results.append({
            'Desde': start_d,
            'Hasta': (end_d - timedelta(days=1)),
            'Días': dias,
            'Capital Base': capital_base,
            'Tasa E.A. Aplicada (%)': ibc * 1.5,
            'Interés Generado en el Periodo': int_gen,
            'Abono a Intereses': abono_interes_periodo,
            'Abono a Capital': abono_capital_periodo,
            'Saldo Capital Acumulado': capital_base,
            'Saldo Intereses Acumulados': intereses_acumulados,
            'Total Fila (Capital + Intereses)': capital_base + intereses_acumulados
        })
        
    df_res = pd.DataFrame(results)
    
    # KPIs
    st.markdown("### Consolidado Final")
    k1, k2, k3, k4 = st.columns(4)
    sf_cap = df_res['Saldo Capital Acumulado'].iloc[-1] if not df_res.empty else 0.0
    sf_int = df_res['Saldo Intereses Acumulados'].iloc[-1] if not df_res.empty else 0.0
    gt = sf_cap + sf_int + intereses_previos
    
    with k1: st.markdown(f"<div class='metric-card'><div class='metric-label'>Saldo Final Capital</div><div class='metric-value'>${sf_cap:,.2f}</div></div>", unsafe_allow_html=True)
    with k2: st.markdown(f"<div class='metric-card'><div class='metric-label'>Saldo Final Intereses</div><div class='metric-value'>${sf_int:,.2f}</div></div>", unsafe_allow_html=True)
    with k3: st.markdown(f"<div class='metric-card'><div class='metric-label'>Int. Corrientes Previos</div><div class='metric-value'>${intereses_previos:,.2f}</div></div>", unsafe_allow_html=True)
    with k4: st.markdown(f"<div class='metric-card'><div style='border-left:5px solid #00f0ff;'><div class='metric-label'>Gran Total a Pagar</div><div class='metric-value' style='color:#00f0ff;'>${gt:,.2f}</div></div></div>", unsafe_allow_html=True)
    
    st.markdown("### Cuadro de Liquidación")
    st.dataframe(df_res.style.format({
        'Capital Base': '${:,.2f}', 'Interés Generado en el Periodo': '${:,.2f}',
        'Abono a Intereses': '${:,.2f}', 'Abono a Capital': '${:,.2f}',
        'Saldo Capital Acumulado': '${:,.2f}', 'Saldo Intereses Acumulados': '${:,.2f}',
        'Total Fila (Capital + Intereses)': '${:,.2f}',
        'Tasa E.A. Aplicada (%)': '{:.2f}%'
    }), use_container_width=True)
    
    # Export
    totales_res = {
        'Saldo Final Capital': sf_cap,
        'Saldo Final Intereses Moratorios': sf_int,
        'Intereses Corrientes Previos': intereses_previos,
        'Gran Total a Pagar': gt
    }
    
    # Excel Extractor
    def to_excel(df_in, v_cuotas_in, v_abonos_in, fecha_liq, int_previos, texto_constancia, f_nombre, f_cargo, f_img):
        out = io.BytesIO()
        df_export = df_in.copy()
        df_export['Desde'] = df_export['Desde'].astype(str)
        df_export['Hasta'] = df_export['Hasta'].astype(str)
        
        tmp_name = None
        if f_img is not None:
            with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp:
                tmp_name = tmp.name
            f_img.save(tmp_name, format="PNG")
            
        try:
            with pd.ExcelWriter(out, engine='openpyxl') as w:
                sheet_name = 'Liquidación'
                current_row = 0
                
                # Resumen Datos
                if texto_constancia and texto_constancia.strip() != "":
                    pd.DataFrame([[texto_constancia.strip()]]).to_excel(w, sheet_name=sheet_name, index=False, header=False, startrow=current_row)
                    current_row += 3
                    
                pd.DataFrame([["DATOS DILIGENCIADOS"]]).to_excel(w, sheet_name=sheet_name, index=False, header=False, startrow=current_row)
                current_row += 1
                pd.DataFrame([{
                    "Fecha de Liquidación": str(fecha_liq),
                    "Intereses Corrientes Previos": f"${int_previos:,.2f}"
                }]).to_excel(w, sheet_name=sheet_name, index=False, startrow=current_row)
                current_row += 3
                
                if not v_cuotas_in.empty:
                    pd.DataFrame([["CUOTAS DE CAPITAL"]]).to_excel(w, sheet_name=sheet_name, index=False, header=False, startrow=current_row)
                    current_row += 1
                    v_cuotas_in.to_excel(w, sheet_name=sheet_name, index=False, startrow=current_row)
                    current_row += len(v_cuotas_in) + 2
                    
                if not v_abonos_in.empty:
                    pd.DataFrame([["ABONOS REALIZADOS"]]).to_excel(w, sheet_name=sheet_name, index=False, header=False, startrow=current_row)
                    current_row += 1
                    v_abonos_in.to_excel(w, sheet_name=sheet_name, index=False, startrow=current_row)
                    current_row += len(v_abonos_in) + 2
                    
                pd.DataFrame([["TABLA DE LIQUIDACIÓN"]]).to_excel(w, sheet_name=sheet_name, index=False, header=False, startrow=current_row)
                current_row += 1
                df_export.to_excel(w, sheet_name=sheet_name, index=False, startrow=current_row)
                
                current_row += len(df_export) + 3
                
                ws = w.sheets[sheet_name]
                if f_img is not None and tmp_name:
                    img_xl = OpenpyxlImage(tmp_name)
                    # Mantener proporciones base
                    ratio = img_xl.height / img_xl.width if img_xl.width > 0 else 1
                    img_xl.width = 200
                    img_xl.height = int(200 * ratio)
                    ws.add_image(img_xl, f"B{current_row + 1}")
                    current_row += int(img_xl.height / 15) + 1 # Estimar filas ocupadas
                            
                if f_nombre.strip() or f_cargo.strip():
                    if f_nombre.strip():
                        ws.cell(row=current_row + 1, column=2, value=f_nombre.strip().upper())
                        ws.cell(row=current_row + 1, column=2).font = Font(bold=True)
                        current_row += 1
                    if f_cargo.strip():
                        ws.cell(row=current_row + 1, column=2, value=f_cargo.strip())
        finally:
            if tmp_name and os.path.exists(tmp_name):
                try:
                    os.remove(tmp_name)
                except:
                    pass
                    
        return out.getvalue()
        
    def to_pdf(df_in, ts, v_cuotas_in, v_abonos_in, fecha_liq, int_previos, texto_constancia, f_nombre, f_cargo, f_img):
        p = FPDF(orientation='L', format='A4')
        p.add_page()
        p.set_font('Arial', 'B', 14)
        p.cell(0, 10, 'Liquidador de Intereses Moratorios Judiciales', 0, 1, 'C')
        
        if texto_constancia and texto_constancia.strip() != "":
            p.ln(2)
            p.set_font('Arial', '', 10)
            p.multi_cell(0, 5, texto_constancia.strip(), align='J')
            p.ln(4)
        
        # Datos Iniciales
        p.set_font('Arial', 'B', 10)
        p.cell(0, 8, 'Datos Diligenciados', 0, 1, 'L')
        p.set_font('Arial', '', 9)
        p.cell(0, 6, f'Fecha de Liquidacion: {fecha_liq}     Intereses Corrientes Previos: ${int_previos:,.2f}', 0, 1, 'L')
        p.ln(2)
        
        if not v_cuotas_in.empty:
            p.set_font('Arial', 'B', 9)
            p.cell(0, 6, 'Cuotas de Capital', 0, 1, 'L')
            p.set_font('Arial', '', 8)
            for _, r in v_cuotas_in.iterrows():
                detalle = r.get('Detalle', '')
                p.cell(0, 5, f"- Detalle: {detalle}       Capital: ${r['Valor Capital']:,.2f}       Vence: {r['Fecha de Vencimiento']}", 0, 1, 'L')
            p.ln(2)
            
        if not v_abonos_in.empty:
            p.set_font('Arial', 'B', 9)
            p.cell(0, 6, 'Abonos Realizados', 0, 1, 'L')
            p.set_font('Arial', '', 8)
            for _, r in v_abonos_in.iterrows():
                p.cell(0, 5, f"- Abono: ${r['Valor Abono']:,.2f}       Fecha: {r['Fecha Abono']}", 0, 1, 'L')
            p.ln(4)
        
        texto_metodologia = (
            "Metodología de Liquidación: La presente liquidación se rige por los preceptos del Código General del Proceso y la jurisprudencia aplicable. Se respeta la prohibición de anatocismo al mantener el capital base inalterado; los intereses causados se relacionan en una columna independiente de acumulados. La tasa de interés aplicable corresponde a 1.5 veces el Interés Bancario Corriente certificado por la Superintendencia Financiera en su modalidad Efectiva Anual (E.A.). Para cada periodo liquidado, la tasa E.A. es convertida a su equivalencia fraccionada aplicando la fórmula exponencial matemática [(1+EA)^(días/365)-1], garantizando el cálculo exacto de los intereses sin acudir a tasas nominales divididas aritméticamente."
        )
        p.set_font('Arial', '', 8)
        p.multi_cell(0, 4, texto_metodologia, align='J')
        p.ln(2)

        p.set_font('Arial', 'B', 10)
        p.cell(0, 8, 'Tabla de Liquidacion', 0, 1, 'L')
        p.set_font('Arial', size=7)
        cols = ['Desde', 'Hasta', 'Dias', 'Capital', 'Tasa E.A%', 'Int.Gen', 'Abo.Int', 'Abo.Cap', 'SF.Cap', 'SF.Int', 'Total']
        wds = [20, 20, 10, 28, 18, 28, 22, 22, 35, 36, 38]
        for i, c in enumerate(cols):
            p.cell(wds[i], 8, c, border=1, align='C')
        p.ln()
        for _, r in df_in.iterrows():
            p.cell(wds[0], 6, str(r['Desde']), border=1)
            p.cell(wds[1], 6, str(r['Hasta']), border=1)
            p.cell(wds[2], 6, str(r['Días']), border=1, align='C')
            p.cell(wds[3], 6, f"${r['Capital Base']:,.2f}", border=1, align='R')
            p.cell(wds[4], 6, f"{r['Tasa E.A. Aplicada (%)']:.2f}%", border=1, align='C')
            p.cell(wds[5], 6, f"${r['Interés Generado en el Periodo']:,.2f}", border=1, align='R')
            p.cell(wds[6], 6, f"${r['Abono a Intereses']:,.2f}", border=1, align='R')
            p.cell(wds[7], 6, f"${r['Abono a Capital']:,.2f}", border=1, align='R')
            p.cell(wds[8], 6, f"${r['Saldo Capital Acumulado']:,.2f}", border=1, align='R')
            p.cell(wds[9], 6, f"${r['Saldo Intereses Acumulados']:,.2f}", border=1, align='R')
            p.cell(wds[10], 6, f"${r['Total Fila (Capital + Intereses)']:,.2f}", border=1, align='R')
            p.ln()
            
        p.ln(10)
        p.set_font('Arial', 'B', 10)
        p.cell(0, 6, 'Resumen Consolidado', 0, 1, 'L')
        p.set_font('Arial', size=9)
        for k, v in ts.items():
            p.cell(70, 6, str(k), border=1)
            p.cell(40, 6, f"${v:,.2f}", border=1, align='R')
            p.ln()
            
        p.ln(15)
        if p.get_y() > 165:
            p.add_page()
            
        if f_img is not None:
            with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp:
                tmp_name = tmp.name
            f_img.save(tmp_name, format="PNG")
            try:
                p.image(tmp_name, w=40)
            finally:
                if os.path.exists(tmp_name):
                    try:
                        os.remove(tmp_name)
                    except: pass
            p.ln(5)
        elif f_nombre.strip() or f_cargo.strip():
            p.ln(10)
            p.cell(60, 0, '', border='T')
            p.ln(2)
            
        if f_nombre.strip() or f_cargo.strip():
            p.set_font('Arial', 'B', 10)
            p.cell(0, 5, str(f_nombre).strip().upper(), 0, 1, 'L')
            p.set_font('Arial', '', 10)
            p.cell(0, 5, str(f_cargo).strip(), 0, 1, 'L')
            
        return bytes(p.output())
    
    e1, e2, e3 = st.columns([1,1,2])
    with e1:
        st.download_button("📥 Descargar Excel", data=to_excel(df_res, v_cuotas, v_abonos, fecha_liquidacion, intereses_previos, texto_inicial, firma_nombre, firma_cargo, firma_imagen), file_name="liquidacion.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    with e2:
        try:
            pdf_bytes = to_pdf(df_res, totales_res, v_cuotas, v_abonos, fecha_liquidacion, intereses_previos, texto_inicial, firma_nombre, firma_cargo, firma_imagen)
            st.download_button("📄 Descargar PDF", data=pdf_bytes, file_name="liquidacion.pdf", mime="application/pdf")
        except Exception as e:
            st.warning(f"Error generando PDF: {e}")
